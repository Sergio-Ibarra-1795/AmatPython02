# Exercise 1 from Excel Module
# Author: Goz
# Description: Get information from xlsx file with openpyxl

# You need to import openpyxl
from openpyxl import *

# Load workbook in a variable
# You can load files:
# "read_only" loads a spreadsheet in read-only mode allowing
# you to open very large Excel files.
# "data_only" ignores loading formulas and instead loads
# only the resulting values.
my_workbook = load_workbook(R'C:\Users\Sergio\Documents\SIR_Personal_Dell\Other_courses\PythonAmatDell\Amat_python02\class03_1\example.xlsx')

# Printing the variable's type
print('El archivo es de tipo: ', type(my_workbook))

# Printing the worksheet's names
print('Y tiene las siguientes worksheets: ', my_workbook.sheetnames)

# Load active sheetname in a variable
# Use my_workbook.worksheets[n] for other sheetname
my_sheet = my_workbook.active

# Print the active's worksheet title
print('El worksheet activo es', my_sheet.title)


# Read information from one cell
print('El renglon 4 columna 1 tiene', my_sheet.cell(row=4, column=1).value)

print('La celda A4 tiene', my_sheet['A4'].value)


# Read information from a range of cells
for i in range(2,6):
    name = my_sheet.cell(row=i, column=1).value
    lastname = my_sheet.cell(row=i, column=2).value
    print(f'{i-1}. {name} {lastname}')

# Read information from a range of cells
for i in range(2,6):
    name = my_sheet.cell(row=i, column=1).value
    lastname = my_sheet.cell(row=i, column=2).value
    print(f'{i-1}. {name} {lastname}')

for row in my_sheet.values:
    for value in row:
        print(value)

for row in my_sheet.values:
     name = row[0]
     lastname = row[1]
     print(f'{name} {lastname}')

# print(my_sheet['A1:C3'])
# print(my_sheet['A'])
# print(my_sheet[3])
# print(my_sheet['A:B'])
# print(my_sheet['1:3'])
# print('************')


for row in my_sheet.iter_rows(min_row=1,
                               max_row=2,
                               min_col=1,
                               max_col=3):
    print(row)


for row in my_sheet.iter_rows(min_row=1,
                               max_row=2,
                               min_col=1,
                               max_col=3,
                               values_only=True):
    print(row)



for col in my_sheet.iter_cols(min_row=1,
                               max_row=2,
                               min_col=1,
                               max_col=3):
    print(col)


for col in my_sheet.iter_cols(min_row=1,
                               max_row=3,
                               min_col=1,
                               max_col=4,
                               values_only=True):
    print(col)


for row in my_sheet.rows:
    print(row[0].value, row[1].value)


for col in my_sheet.columns:
     print(col[3].value, col[1].value)

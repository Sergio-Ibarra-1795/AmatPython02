# Exercise 3 from Excel Module
# Author: Goz
# Description: Update info from xlsx file with openpyxl

# You need to import openpyxl
from openpyxl import *

# Load workbook
my_workbook = load_workbook(R'C:\Users\Sergio\Documents\SIR_Personal_Dell\Other_courses\PythonAmatDell\Amat_python02\class03_1\company.xlsx')

# Enter sheet
my_sheet = my_workbook['Productos']

# Save info in sheet
products = [['Nombre','Codigo', 'Precio'],
           ['martillo','HKJH9879', 120],
           ['mazo','MM57865', 624],
           ['desarmador','JLLKJL87', 32],
           ['pala','HKJHKJH', 420],
          ]

for i in range(0, len(products)):
    for j in range(0, len(products[i])):
        my_sheet.cell(column=j+1,row=i+1).value = products[i][j]

# Insert a column before the existing column 1
# my_sheet.insert_cols(idx=1)

# Insert 2 columns between column 2 and 3
# my_sheet.insert_cols(idx=3, amount=2)

# Delete the created columns
#my_sheet.delete_cols(idx=3, amount=2)

# Insert a new row in the beginning
# my_sheet.insert_rows(idx=1)

# Insert 5 new rows in the beginning
# my_sheet.insert_rows(idx=1, amount=5)

# Delete the first 4 rows
# my_sheet.delete_rows(idx=1, amount=4)

# Merge a range of cells
#my_sheet.merge_cells('A7:D7')

# Unmerge a range of cells
#my_sheet.unmerge_cells('A7:D7')

# Add sheet
#sheet = my_workbook.create_sheet('Ejecutivos', 2)

# Delete sheet
#companies = my_workbook['Compañías']
#my_workbook.remove(companies)

# Duplicates
#clients = my_workbook["Clientes"]
#my_workbook.copy_worksheet(clients)

# Freeze
#my_sheet.freeze_panes = 'C2'

# Formula
my_sheet['C6'] = '=SUM(C2:C5)'

my_workbook.save('company2.xlsx')

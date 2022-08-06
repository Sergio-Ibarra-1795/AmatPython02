# Exercise 2 from Excel Module
# Author: Goz
# Description: Write information to xlsx file with openpyxl

# You need to import openpyxl
from openpyxl import *

# Create a workbook
my_workbook = Workbook()

# Create sheets
my_sheet1 = my_workbook.active
my_sheet1.title = 'Clientes'

my_sheet2 = my_workbook.create_sheet(title='Productos')
my_sheet3 = my_workbook.create_sheet(title='Compañías')
my_sheet4 = my_workbook.create_sheet(title='Ventas')


my_sheet1['A1'] = 'Clientes'
my_sheet2['A1'] = 'Productos'
my_sheet3['A1'] = 'Compañías'
my_sheet4['A1'] = 'Ventas'

clients = [['Nombre','Apellido paterno', 'Apellido materno','cliente'],
           ['Arquetipo','Gonzalez','Urquitas','excelente'],
           ['Madalena','Flores','Urquitas','bueno'],
           ['Guadalupe','Guadarrama','Mendieto','malo'],
           ['Marcelon','Valle','Benitez','bueno']
          ]

for i in range(0, len(clients)):
    for j in range(0, len(clients[i])):
        my_sheet1.cell(column=j+1,row=i+1).value = clients[i][j]

my_workbook.save('company.xlsx')
